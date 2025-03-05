# WakeWord-End2End-TrainingandEvaluation-System
# WWETES
# 唤醒词端到端训练和评估系统

import os
import random
import subprocess
import sys
import time
from datetime import datetime
from os import path

import numpy as np
import torch
from openpyxl import Workbook

from howl.utils.args_utils import ArgOption, ArgumentParserBuilder
import json

# 进程管理
import signal
import atexit

import logging
from pathlib import Path
from howl.utils.logger import setup_logger, Logger

# 日志配置（从train.py移植的日志系统）
def configure_logger(workspace_name):
    """配置日志系统"""
    # 创建日志目录
    base_workspace_path = Path(os.getcwd()) / f"workspaces/exp_{workspace_name}_res8"
    log_dir = base_workspace_path / "logs"
    os.makedirs(log_dir, exist_ok=True)
    
    # 创建日志文件路径
    timestamp = time.strftime('%Y%m%d%H%M%S')
    log_file = log_dir / f"WWETES_{workspace_name}_{timestamp}.log"

    # 初始化Logger
    Logger.LOGGER = setup_logger(
        name=Logger.NAME,
        level=logging.INFO,
        use_stdout=True,
        log_path=str(log_file)
    )
    
    return log_file

def get_relevant_config(new_env, command, proc, action_type):
    """提取相关的配置信息"""
    # 定义需要保存的关键环境变量(可编辑)
    relevant_keys = [
        'SEED', 'NUM_EPOCHS', 'VOCAB', 'WEIGHT_DECAY', 'LEARNING_RATE',
        'LR_DECAY', 'BATCH_SIZE', 'MAX_WINDOW_SIZE_SECONDS',
        'EVAL_WINDOW_SIZE_SECONDS', 'EVAL_STRIDE_SIZE_SECONDS',
        'USE_NOISE_DATASET', 'NUM_MELS', 'INFERENCE_SEQUENCE',
        'CUDA_VISIBLE_DEVICES'
    ]
    
    # 只提取相关的环境变量
    relevant_env = {k: new_env.get(k) for k in relevant_keys if k in new_env}
    
    # 构建配置字典
    config = {
        'type': action_type,
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'process_id': proc.pid,
        'command': command,
        'environment': relevant_env,
        'device': f"cuda:{new_env.get('CUDA_VISIBLE_DEVICES', 'N/A')}"
    }
    
    return config

def save_config_to_file(config, workspace_path, env, action):
    """保存配置到文件"""
    import json
    
    # 生成文件名
    filename = f"model-{env.get('SEED', 'N/A')}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{action}_config.json"
    filepath = os.path.join(workspace_path, filename)
    
    # 保存为JSON格式
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4, ensure_ascii=False)
    
    Logger.info(f"配置已保存到: {filepath}")

# 保存所有创建的子进程
child_processes = []
# 在程序退出时清理所有子进程
def cleanup_processes():
    for proc in child_processes:
        try:
            proc.terminate()  # 尝试优雅终止
            proc.wait(timeout=3)  # 等待进程终止
        except:
            try:
                proc.kill()  # 强制终止
            except:
                pass  # 忽略可能的错误
# 注册退出函数
atexit.register(cleanup_processes)

# 处理终止信号
signal.signal(signal.SIGTERM, lambda sig, frame: cleanup_processes())
signal.signal(signal.SIGINT, lambda sig, frame: cleanup_processes())

def is_job_running(grep_command, count_command):
    """Checks if a command is running"""
    # 使用ps aux | grep 命令查找正在运行的进程，并得到输出
    # ps aux 指令用于显示当前系统中所有正在运行的进程及其详细信息。通过 ps aux，你可以查看系统中所有进程的状态，包括进程 ID、用户、CPU 和内存使用情况、启动时间、命令等信息。
    out = subprocess.check_output("ps aux | grep " + grep_command, shell=True)
    # 依据输出计算正在运行的进程数量，计数查找关键字由count_command指定
    num_proc = out.decode("utf-8").count(count_command)
    return num_proc > 0

# 批量运行给定的一组命令，并定期检查每个进程的状态，以便在 GPU 可用时调度下一个作业
def run_batch_commands(commands, envs, workspace_path, action, grep_command="training.run.train", count_command="python -m training.run.train", gpu_id=None):
    # grep_command：这是用于 ps aux | grep 命令的字符串，用于查找正在运行的进程。默认值是 "training.run.train"。
    # count_command：这是用于计数的字符串，用于确定有多少个匹配的进程正在运行。默认值是 "python -m training.run.train"。
    """
    run given set of commands with the corresponding environments
    check the status of each process regularly and schedule the next job whenever GPU is availble
    
    args help:
        gpu_id: 指定的GPU ID（如果为None则使用多GPU调度）
    """
    # pylint: disable=subprocess-popen-preexec-fn
    # 检查可用的 GPU 数量
    num_gpu = torch.cuda.device_count()
    Logger.info(f"检测到 {num_gpu} 个可用的 GPU")

    # 如果指定了GPU ID，记录日志
    if gpu_id is not None:
        Logger.info(f"所有进程将在GPU {gpu_id} 上运行")

    # 每 10 分钟检查一次
    check_up_delay = 600  # check every 10 mins
    num_running_jobs = 0
    sleep_counter = 0

    # 将两个列表 commands 和 envs 逐对组合在一起，以便在循环中同时迭代这两个列表的元素
    for (command, env) in zip(commands, envs):
        for env_key, env_val in env.items():
            os.environ[env_key] = env_val

        # 如果当前正在运行的进程数量等于 GPU 数量，则等待，直到有 GPU 可用
        # 如果指定了 GPU ID，则不等待，所有进程将在指定的 GPU 上并行运行
        if gpu_id is None and num_running_jobs == num_gpu:
            sleep_counter = 0
            while is_job_running(grep_command, count_command):
                sleep_counter += 1
                Logger.info(f"有一些作业正在运行; 等待 {sleep_counter * check_up_delay / 60} 分钟")
                sys.stdout.flush()
                sys.stderr.flush()
                time.sleep(check_up_delay)
            num_running_jobs = 0

        new_env = os.environ.copy()
        
        # 根据是否指定GPU ID来设置CUDA_VISIBLE_DEVICES
        if gpu_id is not None:
            # 如果指定了GPU ID，则所有进程都使用该GPU
            new_env["CUDA_VISIBLE_DEVICES"] = str(gpu_id)
        else:
            # 否则使用原始的多GPU调度逻辑
            new_env["CUDA_VISIBLE_DEVICES"] = str(num_running_jobs)
        
        # 限定可用GPU
        # new_env["CUDA_VISIBLE_DEVICES"] = str(num_running_jobs)
        # python中的subprocess详解：https://blog.csdn.net/super_he_pi/article/details/99713374
        proc = subprocess.Popen(command.split(), preexec_fn=os.setpgrp, env=new_env)
        child_processes.append(proc) # 将创建的子进程保存到子进程管理池中
        # 参数：子进程组
        # preexec_fn=os.setpgrp：在子进程执行之前调用 os.setpgrp 函数。这会创建一个新的进程组，
        # 并将子进程加入该组。这样做的好处是可以更好地控制子进程，特别是在需要终止整个进程组时。
        # 注意：preexec_fn=os.setpgrp 的作用是创建一个新的进程组，并将子进程设置为该进程组的组长。这会导致几个重要的后果：
        # 1. 进程组独立性：子进程被放入一个新的进程组，与父进程隔离
        # 2. 信号传递隔离：当你使用普通的 kill 命令终止父进程时，该信号不会自动传递给新进程组中的子进程
        # 3. 守护进程行为：子进程成为了独立的进程组，不再依赖于父进程的生命周期
        # 参数：env
        # 如果env不是None，则子程序的环境变量由env的值来设置，而不是默认那样继承父进程的环境变量。
        # 注意，即使你只在env里定义了某一个环境变量的值，也会阻止子程序得到其他的父进程的环境变量
        if action == "train":
            Logger.info("Training model with the following configurations:")
        elif action == "eval":
            Logger.info("Evaluating model with the following configurations:")
        Logger.info("Process ID: {}".format(proc.pid))
        Logger.info("Command: {}".format(command))
        Logger.info("Environment Variables:")
        Logger.info(new_env)
        # flush=True 的作用是强制刷新输出缓冲区。这意味着打印的内容会立即输出到控制台，
        # 而不是等待缓冲区满了之后再输出。这在需要实时查看输出结果时非常有用，特别是在长时间运行的脚本中。
        
        # 不仅在控制台输出一遍参数配置及环境变量，同时将模型信息、参数配置及环境变量保存到workspace_path文件夹中
        # 保存配置信息到文件
        # 保存配置信息
        # if action == "train":
        #     config = get_relevant_config(new_env, command, proc, "Training")
        #     save_config_to_file(config, workspace_path, env, "train") # workspace_path配置有误，从传入run_batch_commands函数时就错了，默认保存到最后一个创建的模型的文件夹中
        # elif action == "eval":
        #     config = get_relevant_config(new_env, command, proc, "Evaluation")
        #     save_config_to_file(config, workspace_path, env, "eval") # workspace_path配置有误，从传入run_batch_commands函数时就错了，默认保存到最后一个创建的模型的文件夹中
            
            
        time.sleep(60)  # add some delay between each job scheduling
        if gpu_id is None:
            num_running_jobs += 1

    sleep_counter = 0
    while is_job_running(grep_command, count_command):
        sleep_counter += 1
        Logger.info(f"有一些作业正在运行; 等待 {sleep_counter * check_up_delay / 60} 分钟")
        time.sleep(check_up_delay)

def main():
    """
    This script uses every GPU on the machine to train multiple res8 models for user-defined
    evaluates them with different threashold, and generate performance reports

    Two different performance reports will be generated, one with the clean audio and one with audios with noise
    
    command help:
    # experiment settings
    --workspace_name: type of the experiment, hey_fire_fox or hey_snips
    --num_models: number of models to train and evaluate (model's name will be generated with random seeds)
    --seed: random seed for the experiment
    --use_given_model: whether to use given model to train or evaluate
    --given_model_seeds: if args.use_given_model is true, given model seeds to train or evaluate
    --exp_type: experiment type
    # training settings
    --num_epochs: number of epochs for training
    --vocab: vocabulary for the model
    --weight_decay: weight decay for the optimizer
    --learning_rate: learning rate for the optimizer
    --lr_decay: learning rate decay for the optimizer
    --batch_size: batch size for training
    --max_window_size_seconds: maximum window size for the training
    --eval_window_size_seconds: window size for evaluation
    --eval_stride_size_seconds: stride size for evaluation
    --use_noise_dataset: whether to use noise dataset
    --num_mels: number of mel bands
    --eval-freq: frequency of evaluation
    --device: device to run the training
    # inference settings
    --inference_sequence: sequence of the words to be detected
    --hop_size: hop size for threshold, if 0.05, thresholds will be [0, 0.05, 0.1, 0.15, ..., 1]
    # dataset settings
    # --dataset_path: path to the dataset(not used in this script)
    --positive_dataset_path: path to the positive dataset
    --negative_dataset_path: path to the negative dataset
    --noise_dataset_path: path to the noise dataset
    --use-stitched-datasets: whether to use stitched datasets, default is false
    
    sample command:
    exp1: hey_fire_fox
    python -m training.run.eval_wake_word_detection --num_models 3 --seed 114514 --workspace_name "hey_fire_fox" --num_epochs "300" --vocab '["hey","fire","fox"]' --weight_decay "0.00001" --learning_rate "0.01" --lr_decay "0.98" --batch_size "16" --max_window_size_seconds "0.5" --eval_window_size_seconds "0.5" --eval_stride_size_seconds "0.063" --use_noise_dataset "True" --num_mels "40" --eval-freq "20" --device "cuda:1" --inference_sequence "[0,1,2]" --hop_size 0.05 --positive_dataset_path "/home/hrq/DHG-Workspace/Research_on_Low-Cost_Custom_Voice_Wake-Up_Based_on_Voice_Cloning/baselines/KWS/howl/datasets/hey_fire_fox/positive" --negative_dataset_path "/home/hrq/DHG-Workspace/Research_on_Low-Cost_Custom_Voice_Wake-Up_Based_on_Voice_Cloning/baselines/KWS/howl/datasets/hey_fire_fox/negative" --noise_dataset_path "/home/hrq/DHG-Workspace/Research_on_Low-Cost_Custom_Voice_Wake-Up_Based_on_Voice_Cloning/datasets/Hey-Fire-Fox/hey-ff-noise/MS-SNSD"

    train 10 models for hey firefox with random seeds using datasets x and y
    then evaluate them for thresholds ranging from 0 to 1 in increments of 0.05
    """
    # pylint: disable=too-many-branches
    # pylint: disable=too-many-statements
    apb = ArgumentParserBuilder()
    apb.add_options(
        # experiment settings
        ArgOption("--workspace_name", type=str, choices=["hey_fire_fox", "hey_snips"], default="hey_fire_fox"), # 现在workspace_name仅作为设置workspace的文件夹名称的依据，可以自定义为任何名称
        ArgOption("--num_models", type=int, default=1, help="number of models to train and evaluate"),
        ArgOption("--seed", type=int, default=0), # 用作生成不同模型id的random seed
        ArgOption("--use_given_model", action="store_true", help="use the given model to train or eval"), # 使用使用指定seed的模型
        ArgOption("--given_model_seeds", type=str, default='[]'), # 输入示例：'["123","456","789"]'
        ArgOption("--exp_type", type=str, default='train-eval'), # 指定需要进行的实验类型
        # 实验类型包括：
            # 1. train-eval：先进行模型训练，后进行模型评估
            # 2. train：只进行模型训练
            # 3. eval：只进行模型评估
            # 4. fine-tuning：继续训练（微调），将执行含有--load-weights参数的模型训练，不含评估操作
            # 5. reports：生成报告，不含训练和评估操作
        # training settings
        ArgOption("--num_epochs", type=str, default="300"),
        ArgOption("--vocab", type=str, default='["hey","fire","fox"]'),
        ArgOption("--weight_decay", type=str, default="0.00001"),
        ArgOption("--learning_rate", type=str, default="0.01"),
        ArgOption("--lr_decay", type=str, default="0.98"),
        ArgOption("--batch_size", type=str, default="16"),
        ArgOption("--max_window_size_seconds", type=str, default="0.5"),
        ArgOption("--eval_window_size_seconds", type=str, default="0.5"),
        ArgOption("--eval_stride_size_seconds", type=str, default="0.063"),
        ArgOption("--use_noise_dataset", type=str, default="True"),
        ArgOption("--num_mels", type=str, default="40"),
        ArgOption("--eval-freq", type=str, default="20"),
        ArgOption("--device", type=str, default=""),  
        # 如果device参数以"cuda:"开头，则将其解析为GPU ID，且仅使用指定的GPU ID运行训练和评估，所有并行进程都将在该GPU上运行
        # 若不指定，则使用多GPU调度，并行进程数<=设备可用gpu数量
        # 默认值为 ""，即若不指定device，默认使用多GPU调度
        # inference settings
        ArgOption("--inference_sequence", type=str, default="[0,1]"),
        ArgOption("--hop_size", type=float, default=0.05, help="hop size for threshold"),
        # dataset settings
        # ArgOption("--dataset_path", type=str, default="/data/speaker-id-split-medium"),
        ArgOption("--positive_dataset_path", type=str, default="/data/positive_dataset"),
        ArgOption("--negative_dataset_path", type=str, default="/data/negative_dataset"),
        ArgOption("--noise_dataset_path", type=str, default="/data/MS-SNSD"),
        ArgOption("--use-stitched-datasets", action="store_true"),
        # 对于参数--use-stitched-datasets。参数的类型是 bool（布尔类型）。当提供这个参数时，args.use_stitched_datasets 的值将是 True。
        # 如果你不提供这个参数，args.use_stitched_datasets 的值将保持为 False。
    )

    args = apb.parser.parse_args()
    # [参数传递中中划线和下划线的转换]
    # 在使用 argparse 解析命令行参数时，命令行参数中的中划线 (-) 
    # 会自动转换为下划线 (_) 以匹配 Python 变量命名的惯例。
    # 这意味着在命令行中使用 --eval-freq 参数时，解析后的结果会存储在
    # args.eval_freq 中。
    # 对于原本带有下划线的参数，例如 --batch_size，
    # 它们会直接转换为带有下划线的变量名 args.batch_size
    
    # 设置日志系统
    log_file = configure_logger(args.workspace_name)
    Logger.heading("=============== 初始化日志系统 ===============")
    Logger.info(f"日志文件创建于: {log_file}")
    Logger.info(f"实验类型: {args.exp_type}")
    Logger.info(f"工作空间名称: {args.workspace_name}")
    # 记录重要参数
    Logger.heading("=============== 设置实验参数 ===============")
    Logger.info(f"模型数量: {args.num_models}")
    Logger.info(f"随机种子: {args.seed}")
    Logger.info(f"是否使用指定模型: {args.use_given_model}")
    if args.use_given_model:
        Logger.info(f"指定模型种子: {args.given_model_seeds}")
    Logger.info(f"阈值步长: {args.hop_size}")
    
    # TODO: 提取指定GPU的ID
    gpu_id = 5
    if args.device and args.device.startswith("cuda:"):
        try:
            gpu_id = int(args.device.split(":")[1])
            Logger.info(f"将使用指定的GPU: {gpu_id}")
        except (IndexError, ValueError):
            Logger.warning(f"无法解析设备参数: {args.device}，将使用多GPU调度")
    
    Logger.info(f"是否使用拼接数据集: {args.use_stitched_datasets}")
    Logger.info(f"数据集路径: 正样本-{args.positive_dataset_path}, 负样本-{args.negative_dataset_path}")
    Logger.info(f"实验类型: {args.exp_type}")

    # 设置随机种子
    random.seed(args.seed)

    # total sample counts, these number can be found when the datasets are loaded for training
    # these numbers are used to calculate the sum of the metrics

    ######### deprecated code #########
    # if args.workspace_name == "hey_fire_fox":
    #     total_counts = {"Dev positive": "76", "Dev negative": "2531", "Test positive": "54", "Test negative": "2504"}
    # elif args.workspace_name == "hey_snips":
    #     total_counts = {
    #         "Dev positive": "2484",
    #         "Dev negative": "13598",
    #         "Test positive": "2529",
    #         "Test negative": "13943",
    #     }
    ######### deprecated code #########
    
    # 不使用静态设置的数据集大小，使用动态计算的数据集大小
    # 统计数据集中的正例和负例数量
    if args.use_stitched_datasets:
        dev_positive_path = os.path.join(args.positive_dataset_path, "stitched-metadata-dev.jsonl")
        test_positive_path = os.path.join(args.positive_dataset_path, "stitched-metadata-test.jsonl")
        dev_negative_path = os.path.join(args.negative_dataset_path, "stitched-metadata-dev.jsonl")
        test_negative_path = os.path.join(args.negative_dataset_path, "stitched-metadata-test.jsonl")
    else:
        dev_positive_path = os.path.join(args.positive_dataset_path, "aligned-metadata-dev.jsonl")
        test_positive_path = os.path.join(args.positive_dataset_path, "aligned-metadata-test.jsonl")
        dev_negative_path = os.path.join(args.negative_dataset_path, "aligned-metadata-dev.jsonl")
        test_negative_path = os.path.join(args.negative_dataset_path, "aligned-metadata-test.jsonl")
    
    def count_lines_in_file(filepath):
        with open(filepath, 'r') as file:
            return sum(1 for _ in file)
    
    total_counts = {
        "Dev positive": str(count_lines_in_file(dev_positive_path)),
        "Dev negative": str(count_lines_in_file(dev_negative_path)),
        "Test positive": str(count_lines_in_file(test_positive_path)),
        "Test negative": str(count_lines_in_file(test_negative_path)),
    }
    
    # 输出统计信息
    Logger.heading("=============== 数据集统计 ===============")
    for key, value in total_counts.items():
        Logger.info(f"{key}: {value}")

    # dataset settings
    # os.environ["DATASET_PATH"] = args.dataset_path
    os.environ["POSITIVE_DATASET_PATH"] = args.positive_dataset_path
    os.environ["NEGATIVE_DATASET_PATH"] = args.negative_dataset_path
    os.environ["NOISE_DATASET_PATH"] = args.noise_dataset_path
    # Training settings
    os.environ["NUM_EPOCHS"] = args.num_epochs
    os.environ["VOCAB"] = args.vocab
    os.environ["WEIGHT_DECAY"] = args.weight_decay
    os.environ["LEARNING_RATE"] = args.learning_rate
    os.environ["LR_DECAY"] = args.lr_decay
    os.environ["BATCH_SIZE"] = args.batch_size
    os.environ["MAX_WINDOW_SIZE_SECONDS"] = args.max_window_size_seconds
    os.environ["EVAL_WINDOW_SIZE_SECONDS"] = args.eval_window_size_seconds
    os.environ["EVAL_STRIDE_SIZE_SECONDS"] = args.eval_stride_size_seconds
    os.environ["USE_NOISE_DATASET"] = str(args.use_noise_dataset.lower() == "true") # 将字符串转换为小写，然后检查是否等于 "true"
    os.environ["NUM_MELS"] = args.num_mels
    # Inference settings
    os.environ["INFERENCE_SEQUENCE"] = args.inference_sequence

    Logger.heading("=============== 模型seeds生成 ===============")

    seeds = []
    def get_workspace_path(workspace_name, seed):
        return os.getcwd() + f"/workspaces/exp_{workspace_name}_res8/" + str(seed)

    if args.use_given_model:
        Logger.info("====== use given model in seed list ======")
        # 使用json.loads()解析JSON格式的字符串为Python列表
        seeds = json.loads(args.given_model_seeds)
        # 确保列表中的所有元素都是字符串类型
        seeds = [str(seed) for seed in seeds]
        Logger.info(f"使用指定的{len(seeds)}个模型种子: {seeds}")
    else:
        Logger.info(f"====== use random generated models, train and eval {args.num_models} models ======")
        for _ in range(args.num_models):
            seed = str(random.randint(1, 1000000))
            seeds.append(seed)
        Logger.info(f"生成{args.num_models}个随机模型种子: {seeds}")

    # generate commands to run along with the environments
    # 根据实验类型生成命令
    if args.exp_type in ['train-eval', 'train', 'fine-tuning']:
        Logger.heading("=============== training each models ===============")
        training_commands = []
        training_envs = []
        for seed in seeds:
            env = {}
            # env = os.environ.copy()  # 复制当前全局的环境变量
            env["SEED"] = seed
            # 上面新设置的env环境最终也是会在run_batch_commands函数中被使用，通过os.environ[env_key] = env_val被设置到全局环境变量中
            # 和上面设置全局环境变量的方式一样，最后都会汇总到全局环境变量中
            # in run_batch_commands():
                # ...
                # for env_key, env_val in env.items():
                #     os.environ[env_key] = env_val
                # ...
            workspace_path = get_workspace_path(args.workspace_name, seed)
            os.system("mkdir -p " + workspace_path)
            if args.use_stitched_datasets:
                command = (
                    "python -m training.run.train --model res8 --workspace " + workspace_path 
                    + " -i " + args.positive_dataset_path + " " + args.negative_dataset_path 
                    + " --eval-freq " + args.eval_freq
                    + " --eval-hop-size " + str(args.hop_size)
                    + " --device "  + "cuda:0"
                    + " --use-stitched-datasets"
                )
            else:
                command = (
                    "python -m training.run.train --model res8 --workspace " + workspace_path 
                    + " -i " + args.positive_dataset_path + " " + args.negative_dataset_path 
                    + " --eval-freq " + args.eval_freq
                    + " --eval-hop-size " + str(args.hop_size)
                    + " --device "  + "cuda:0"
                )
            training_commands.append(command)
            training_envs.append(env)

        # 执行训练命令
        if training_commands:
        # 在 Python 中，空列表（[]）在布尔上下文中被视为 False，而非空列表则被视为 True
            Logger.info(f"开始执行{len(training_commands)}个训练命令...")
            run_batch_commands(training_commands, training_envs, workspace_path, action="train", gpu_id=gpu_id)
        else:
            Logger.info("没有训练命令需要执行，请检查生成的训练命令！")

    # 根据实验类型生成评估命令
    if args.exp_type in ['train-eval', 'eval']:
        Logger.heading("=============== evaluating each models ===============")
        eval_commands = []
        eval_envs = []
        
        for seed in seeds:
            env = {}
            env["SEED"] = seed
            workspace_path = get_workspace_path(args.workspace_name, seed)
            if args.use_stitched_datasets:
                command = (
                    "python -m training.run.train --eval --model res8 --workspace " + workspace_path
                    + " -i " + args.positive_dataset_path + " " + args.negative_dataset_path
                    + " --eval-hop-size " + str(args.hop_size)
                    + " --device "  + "cuda:0"
                    + " --use-stitched-datasets"
                )
            else:
                command = (
                    "python -m training.run.train --eval --model res8 --workspace " + workspace_path
                    + " -i " + args.positive_dataset_path + " " + args.negative_dataset_path
                    + " --eval-hop-size " + str(args.hop_size)
                    + " --device "  + "cuda:0"
                )
                
            eval_commands.append(command)
            eval_envs.append(env)

        if eval_commands:
            # 在 Python 中，空列表（[]）在布尔上下文中被视为 False，而非空列表则被视为 True
            Logger.info(f"开始执行{len(eval_commands)}个评估命令...")
            run_batch_commands(eval_commands, eval_envs, workspace_path, action="eval", gpu_id=gpu_id)
        else:
            Logger.info("没有评估命令需要执行，请检查生成的评估命令！")

    # generate reports
    # 在训练及评估、仅评估、生成报告的选项下生成报告及ROC曲线
    if args.exp_type in ['train-eval', 'eval', 'reports']:
        Logger.heading("=============== generating reports ===============")
        
        # 获取当前时间戳(开始生成报告的时间戳)
        now = datetime.now()
        dt_string = now.strftime("%Y%m%d%H%M%S")
        # 创建工作簿
        clean_wb = Workbook()
        noisy_wb = Workbook()
        clean_sheets = {}
        noisy_sheets = {}
        # 初始化报告
        col_mapping = {
            "Dev positive": "B",
            "Dev noisy positive": "B",
            "Dev negative": "H",
            "Dev noisy negative": "H",
            "Test positive": "N",
            "Test noisy positive": "N",
            "Test negative": "T",
            "Test noisy negative": "T",
        }
        metrics = ["threshold", "tp", "tn", "fp", "fn"]
        clean_col_names = ["Dev positive", "Dev negative", "Test positive", "Test negative"]
        # 计算阈值列表
        def round_and_convert_to_str(num):
            return str(round(num, 2))
        raw_thresholds = np.arange(0, 1.000001, args.hop_size)
        thresholds = list(map(round_and_convert_to_str, raw_thresholds))
        # 聚合指标计算
        clean_cols_aggregated = {}
        noisy_cols_aggregated = {}
        # 计算聚合指标
        def compute_aggregated_metrics(sheet, col_idx, results):
            sheet[col_idx + "3"] = str(results.mean())
            sheet[col_idx + "4"] = str(results.std())
            sheet[col_idx + "5"] = str(np.percentile(results, 90))
            sheet[col_idx + "6"] = str(np.percentile(results, 95))
            sheet[col_idx + "7"] = str(np.percentile(results, 99))
            sheet[col_idx + "8"] = str(results.sum())
        # 函数 get_cell_idx() 用于获取单元格的索引
        def get_cell_idx(char, ind):
            return chr(ord(char) + ind)
        # 函数 prepare_report() 用于准备报告
        def prepare_report(sheet):
            sheet["A3"] = "mean"
            sheet["A4"] = "std"
            sheet["A5"] = "p90"
            sheet["A6"] = "p95"
            sheet["A7"] = "p99"
            sheet["A8"] = "sum"
            
            for col in clean_col_names:
                cell_idx = col_mapping[col] + "1"
                sheet[cell_idx] = col
                cell_idx = get_cell_idx(col_mapping[col], 1) + "1"
                sheet[cell_idx] = total_counts[col]
                
                for i, metric in enumerate(metrics):
                    cell_idx = get_cell_idx(col_mapping[col], i) + "2"
                    sheet[cell_idx] = metric
        
        # 为每个阈值创建一个xlsx子工作表
        for idx, threshold in enumerate(thresholds):
            clean_sheets[threshold] = clean_wb.create_sheet(threshold, idx)
            prepare_report(clean_sheets[threshold])
            noisy_sheets[threshold] = noisy_wb.create_sheet(threshold, idx)
            prepare_report(noisy_sheets[threshold])

            for col_name in clean_col_names:
                for metric_idx, metric in enumerate(metrics):
                    target_metric = threshold + "_" + get_cell_idx(col_mapping[col_name], metric_idx)
                    clean_cols_aggregated[target_metric] = []
                    noisy_cols_aggregated[target_metric] = []

        # 创建报告文件夹
        base_workspace_path = os.getcwd() + f"/workspaces/exp_{args.workspace_name}_res8"
        report_dir = f"{base_workspace_path}/exp_results"
        os.system(f"mkdir -p {report_dir}")
        
        # 命名报告命名（含精确到秒的时间戳）
        clean_file_path = f"{report_dir}/{args.workspace_name}_clean_{dt_string}.xlsx"
        noisy_file_path = f"{report_dir}/{args.workspace_name}_noisy_{dt_string}.xlsx"


        for seed_idx, seed in enumerate(seeds):
            for _, threshold in enumerate(thresholds):
                clean_sheet = clean_sheets[threshold]
                noisy_sheet = noisy_sheets[threshold]

                row = str(seed_idx + 10)
                clean_sheet["A" + row] = seed
                noisy_sheet["A" + row] = seed

                workspace_path = get_workspace_path(args.workspace_name, seed)
                result_path = workspace_path + "/" + threshold + "_results.csv"
                # 代码使用tail -n 8读取结果文件最后8行，然后将结果拆分为行
                raw_result = subprocess.check_output(["tail", "-n", "8", result_path]).decode("utf-8")
                results = raw_result.split("\n")

                # parse and update the report
                for result in results:
                    if len(result) == 0:
                        break
                    vals = result.split(",")
                    # 对于每一行记录，代码通过第一列的数据集名称来决定如何处理该行数据
                    key = vals[0] # 提取第一列作为数据集名称
                    start_col = col_mapping[key] # 使用预定义的映射查找对应的起始列
                    # CSV文件中行的顺序可以随意排列，只要满足：
                    # 1. 每行的第一列是有效的数据集名称（如"Dev positive"、"Dev negative"等）
                    # 2. 后续列的数据按照 threshold、tp、tn、fp、fn 的顺序排列
                    # 3. 所有8种数据集类型都存在于文件中
                    # 无论CSV行的顺序如何，只要每行内部格式正确，报告生成逻辑都能正常工作。
                    
                    # 然后根据数据集名称是否包含"noisy"，将数据写入不同的报告表格
                    if "noisy" in key:
                        for metric_idx, metric in enumerate(vals[1:]):
                            col_idx = get_cell_idx(start_col, metric_idx)
                            cell_ind = col_idx + str(row)
                            noisy_sheet[cell_ind] = metric

                            target_metric = threshold + "_" + col_idx
                            try:
                                noisy_cols_aggregated[target_metric].append(float(metric))
                                compute_aggregated_metrics(noisy_sheet, col_idx, np.array(noisy_cols_aggregated[target_metric]))
                            except ValueError:
                                Logger.warning(f"无法将指标 {metric} 转换为浮点数，已跳过 (seed: {seed}, threshold: {threshold}, key: {key})")
                    else:
                        for metric_idx, metric in enumerate(vals[1:]):
                            col_idx = get_cell_idx(start_col, metric_idx)
                            cell_ind = col_idx + str(row)
                            clean_sheet[cell_ind] = metric

                            target_metric = threshold + "_" + col_idx
                            try:
                                clean_cols_aggregated[target_metric].append(float(metric))
                                compute_aggregated_metrics(clean_sheet, col_idx, np.array(clean_cols_aggregated[target_metric]))
                            except ValueError:
                                Logger.warning(f"无法将指标 {metric} 转换为浮点数，已跳过 (seed: {seed}, threshold: {threshold}, key: {key})")
        # 在循环外部最后保存一次报告文件
        
        # openpyxl库的默认行为：创建一个新的Workbook对象时，它会自动生成一个名为"Sheet"的默认工作表。
        # 删除默认的"Sheet"工作表
        if "Sheet" in clean_wb.sheetnames:
            clean_wb.remove(clean_wb["Sheet"])
        if "Sheet" in noisy_wb.sheetnames:
            noisy_wb.remove(noisy_wb["Sheet"])
        
        clean_wb.save(clean_file_path)
        noisy_wb.save(noisy_file_path)
        
        Logger.info(f"\treport for clean setting is generated at {clean_file_path}")
        Logger.info(f"\treport for noisy setting is generated at {noisy_file_path}")
        Logger.info("=============== reports generation finished ===============")
        
        # 调用generate_roc.py生成ROC曲线
        Logger.heading("=============== generating ROC curves ===============")
        try:
            # 提取xlsx文件名
            clean_report_basename = os.path.basename(clean_file_path)
            noisy_report_basename = os.path.basename(noisy_file_path)
            
            # 构建generate_roc命令
            roc_cmd = [
                "python", "-m", "training.run.generate_roc",
                "--workspace_name", base_workspace_path,
                "--clean_report_name", clean_report_basename,
                "--noisy_report_name", noisy_report_basename,
                # TODO: 依据train.py中输出的结果修改dev集和test集的总录音时长，未来可以考虑自动流水线
                "--total_dev_len", "25739.7846",
                "--total_test_len", "24278.9884"
            ]
            
            # 执行命令
            Logger.info(f"执行ROC曲线生成命令: {' '.join(roc_cmd)}")
            result = subprocess.run(roc_cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            
            # 输出命令执行结果
            Logger.info("ROC曲线生成成功！")
            Logger.info(result.stdout.decode('utf-8'))
            
            # 如果有警告或错误信息
            if result.stderr:
                stderr_output = result.stderr.decode('utf-8')
                if stderr_output.strip():
                    Logger.warning("ROC曲线生成过程中的警告或错误信息:")
                    Logger.warning(stderr_output)
                    
        except subprocess.CalledProcessError as e:
            Logger.error(f"ROC曲线生成失败: {e}")
            Logger.error(f"错误详情: {e.stderr.decode('utf-8') if e.stderr else '无详细信息'}")
        except Exception as e:
            Logger.error(f"调用ROC曲线生成时发生错误: {str(e)}")
            
        Logger.info("=============== ROC curves generation finished ===============")

if __name__ == "__main__":
    main()
