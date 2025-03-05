import matplotlib.pyplot as plt
from openpyxl import load_workbook
from datetime import datetime
import os

from howl.utils.args_utils import ArgOption, ArgumentParserBuilder


# 初步实验观察，生成的ROC曲线不是一直向右下的，反而还有些往左上角往回拐，向”打结”一样
# 这可能是由于测试样本不足导致的统计波动，对noisy数据集来说，这种情况更明显。
# 故按FAPH对点进行排序后再绘制曲线：
def plot_sorted_roc(faph, frr, color, label, is_dev=False):
    """按FAPH排序后绘制ROC曲线
    
    Args:
        faph: False Alarm Per Hour 数据
        frr: False Rejection Rate 数据
        color: 曲线颜色
        label: 图例标签
        is_dev: 是否为开发集数据（True为开发集使用虚线，False为测试集使用实线）
    """
    # 删除最后一个点(通常是极端值)
    faph, frr = faph[:-1], frr[:-1]
    
    # 创建点的索引
    points = [(x, y) for x, y in zip(faph, frr)]
    # 按FAPH排序
    points.sort(key=lambda p: p[0])
    
    # 提取排序后的坐标
    sorted_faph = [p[0] for p in points]
    sorted_frr = [p[1] for p in points]
    
    # 根据数据集类型选择线型
    linestyle = "--+" if is_dev else "-+"
    
    # 绘制排序后的曲线
    return plt.plot(sorted_faph, sorted_frr, linestyle, color=color, label=label)

# 或者，也可以只保留单调递减的点，即随FAPH增加，FRR减少或不变的点：
def plot_monotonic_roc(faph, frr, color, label, is_dev=False):
    """按FAPH排序后绘制单调递减的ROC曲线（删除违反单调性的点）
    
    Args:
        faph: False Alarm Per Hour 数据
        frr: False Rejection Rate 数据
        color: 曲线颜色
        label: 图例标签
        is_dev: 是否为开发集数据（True为开发集使用虚线，False为测试集使用实线）
    Returns:
        plot对象和实际绘制的点坐标 (sorted_faph, sorted_frr)
    
    """
    # 删除最后一个点(通常是极端值)
    faph, frr = faph[:-1], frr[:-1]
    
    # 创建点的索引
    points = [(x, y) for x, y in zip(faph, frr)]
    # 按FAPH排序（从小到大）
    points.sort(key=lambda p: p[0])
    
    # 筛选单调递减的点（随FAPH增加，FRR减少或不变）
    monotonic_points = []
    if points:  # 确保有点
        monotonic_points.append(points[0])  # 第一个点总是包含的
        current_min_frr = points[0][1]
        
        for point in points[1:]:
            if point[1] <= current_min_frr:  # 只保留FRR小于等于之前最小值的点
                monotonic_points.append(point)
                current_min_frr = point[1]
    
    # 提取排序后的坐标
    sorted_faph = [p[0] for p in monotonic_points]
    sorted_frr = [p[1] for p in monotonic_points]
    
    # 根据数据集类型选择线型
    linestyle = "--+" if is_dev else "-+"
    
    # 绘制排序后的曲线
    plot = plt.plot(sorted_faph, sorted_frr, linestyle, color=color, label=f"{label}")
    return plot, (sorted_faph, sorted_frr)


def load_metrics(workbook, thresholds, total_dev_len, total_test_len):
    """Reorganize entries in the Excel file into dev_faph, dev_frr, test_faph, test_frr

    Args:
        workbook: Excel工作簿对象
        thresholds: 阈值列表
        total_dev_len: 开发集音频总长度(秒)
        total_test_len: 测试集音频总长度(秒)

    Returns:
        dev_faph, dev_frr, test_faph, test_frr
        # dev_faph: dev_false_alarm_per_hour 即每小时假警报数(假报为正例)，假警报数=FP/(总音频长度/3600)
        # dev_frr: dev_false_rejection_rate 即漏报率，漏报率=FN/(FN+TP)
        # test_faph: test_false_alarm_per_hour 即每小时假警报数(假报为正例)，假警报数=FP/(总音频长度/3600)
        # test_frr: test_false_rejection_rate 即漏报率，漏报率=FN/(FN+TP)
        # 这种表示方式在语音唤醒系统的评估中非常常见，比纯粹的错误率更能代表实际使用体验。
    """

    dev_faph = []
    dev_frr = []

    test_faph = []
    test_frr = []

    for threshold in thresholds:
        sheet = workbook[str(threshold)]

        dev_tp = float(sheet["C3"].value)
        dev_fn = float(sheet["F3"].value)
        # dev_tn = float(sheet["J3"].value)
        dev_fp = float(sheet["K3"].value)

        dev_faph.append(dev_fp / (total_dev_len / 3600))  # per hour metric
        dev_frr.append(dev_fn / (dev_fn + dev_tp))

        test_tp = float(sheet["O3"].value)
        test_fn = float(sheet["R3"].value)
        # test_tn = float(sheet["V3"].value)
        test_fp = float(sheet["W3"].value)

        test_faph.append(test_fp / (total_test_len / 3600))  # per hour metric
        test_frr.append(test_fn / (test_fn + test_tp))

    return dev_faph, dev_frr, test_faph, test_frr


def main():
    """
    This script uses reports generated by run_exp.py to create roc curve

    sample command:
    python -m training.run.generate_roc --exp_timestamp Sep-08-11-28 --exp_type hey_firefox
    """
    apb = ArgumentParserBuilder()
    apb.add_options(
        ArgOption("--workspace_name", type=str, help="工作目录路径", required=True),
        ArgOption("--clean_report_name", type=str, help="干净数据集报告文件名(例如:hey_fire_fox_clean_20250228120442.xlsx)", required=True),
        ArgOption("--noisy_report_name", type=str, help="噪声数据集报告文件名(例如:hey_fire_fox_noisy_20250228120442.xlsx)", required=True),
        ArgOption("--total_dev_len", type=str, default="25739.7846", help="开发集音频总长度(秒)"),
        ArgOption("--total_test_len", type=str, default="24278.9884", help="测试集音频总长度(秒)"),
    )

    args = apb.parser.parse_args()

    # 将输入的字符串转换为浮点数
    total_dev_len = float(args.total_dev_len)
    total_test_len = float(args.total_test_len)
    
    reports_dir = os.path.join(args.workspace_name, "exp_results")
    clean_file_name = os.path.join(reports_dir, args.clean_report_name)
    print("\treport for clean setting is ", clean_file_name)
    clean_wb = load_workbook(clean_file_name)
    
    noisy_file_name = os.path.join(reports_dir, args.noisy_report_name)
    print("\treport for noisy setting is ", noisy_file_name)
    noisy_wb = load_workbook(noisy_file_name)
    
    now = datetime.now()
    dt_string = now.strftime("%Y%m%d%H%M%S")
    save_dir = os.path.join(args.workspace_name, "exp_results")
    os.makedirs(save_dir, exist_ok=True)
    save_path = os.path.join(save_dir, f"roc_{dt_string}.pdf")

    thresholds = []
    for name in clean_wb.sheetnames:
        try:
            thresholds.append(float(name))
        except ValueError:
            print("Not a float: ", name)

    clean_dev_faph, clean_dev_frr, clean_test_faph, clean_test_frr = load_metrics(clean_wb, thresholds, total_dev_len, total_test_len)
    noisy_dev_faph, noisy_dev_frr, noisy_test_faph, noisy_test_frr = load_metrics(noisy_wb, thresholds, total_dev_len, total_test_len)

    # generate plots
    plt.rcParams.update({"font.size": 15})

    plt.xlabel("False Alarms Per Hour")
    plt.ylabel("False Rejection Rate")

    print("thresholds:", thresholds)
    print("clean_dev_faph:", [round(num, 3) for num in clean_dev_faph])
    print("clean_dev_frr:", [round(num, 3) for num in clean_dev_frr])
    print("clean_test_faph:", [round(num, 3) for num in clean_test_faph])
    print("clean_test_frr:", [round(num, 3) for num in clean_test_frr])

    print("noisy_dev_faph:", [round(num, 3) for num in noisy_dev_faph])
    print("noisy_dev_frr:", [round(num, 3) for num in noisy_dev_frr])
    print("noisy_test_faph:", [round(num, 3) for num in noisy_test_faph])
    print("noisy_test_frr:", [round(num, 3) for num in noisy_test_frr])

    # 原始绘制方法
    # plt.plot(clean_dev_faph[:-1], clean_dev_frr[:-1], "--+", color="tab:blue", label="clean dev")
    # plt.plot(clean_test_faph[:-1], clean_test_frr[:-1], "-+", color="tab:blue", label="clean test")
    # plt.plot(noisy_dev_faph[:-1], noisy_dev_frr[:-1], "--+", color="tab:orange", label="noisy dev")
    # plt.plot(noisy_test_faph[:-1], noisy_test_frr[:-1], "-+", color="tab:orange", label="noisy test")

    # 按FAPH对点进行排序后再绘制曲线：
    # plot_sorted_roc(clean_dev_faph, clean_dev_frr, "tab:blue", "clean dev")
    # plot_sorted_roc(clean_test_faph, clean_test_frr, "tab:blue", "clean test (sorted)")
    # plot_sorted_roc(noisy_dev_faph, noisy_dev_frr, "tab:orange", "noisy dev")
    # plot_sorted_roc(noisy_test_faph, noisy_test_frr, "tab:orange", "noisy test (sorted)")
    
    # 使用单调递减的ROC曲线绘制方法
    # plot_monotonic_roc(clean_dev_faph, clean_dev_frr, "tab:blue", "clean dev", is_dev=True)
    # plot_monotonic_roc(clean_test_faph, clean_test_frr, "tab:blue", "clean test", is_dev=False)
    # plot_monotonic_roc(noisy_dev_faph, noisy_dev_frr, "tab:orange", "noisy dev", is_dev=True)
    # plot_monotonic_roc(noisy_test_faph, noisy_test_frr, "tab:orange", "noisy test", is_dev=False)
    _, (clean_dev_points_x, clean_dev_points_y) = plot_monotonic_roc(
        clean_dev_faph, clean_dev_frr, "tab:blue", "clean dev", is_dev=True)
    _, (clean_test_points_x, clean_test_points_y) = plot_monotonic_roc(
        clean_test_faph, clean_test_frr, "tab:blue", "clean test", is_dev=False)
    _, (noisy_dev_points_x, noisy_dev_points_y) = plot_monotonic_roc(
        noisy_dev_faph, noisy_dev_frr, "tab:orange", "noisy dev", is_dev=True)
    _, (noisy_test_points_x, noisy_test_points_y) = plot_monotonic_roc(
        noisy_test_faph, noisy_test_frr, "tab:orange", "noisy test", is_dev=False)


    plt.grid()
    plt.legend()
    plt.savefig(save_path)
    print(f"ROC curve saved to {save_path}")
    
    # 将数据保存到文本文件
    txt_save_path = os.path.join(save_dir, f"roc_{dt_string}.txt")
    with open(txt_save_path, 'w') as f:
        # 写入原始数据
        f.write(f"thresholds: {thresholds}\n")
        f.write(f"clean_dev_faph: {[round(num, 3) for num in clean_dev_faph]}\n")
        f.write(f"clean_dev_frr: {[round(num, 3) for num in clean_dev_frr]}\n")
        f.write(f"clean_test_faph: {[round(num, 3) for num in clean_test_faph]}\n")
        f.write(f"clean_test_frr: {[round(num, 3) for num in clean_test_frr]}\n\n")
        
        f.write(f"noisy_dev_faph: {[round(num, 3) for num in noisy_dev_faph]}\n")
        f.write(f"noisy_dev_frr: {[round(num, 3) for num in noisy_dev_frr]}\n")
        f.write(f"noisy_test_faph: {[round(num, 3) for num in noisy_test_faph]}\n")
        f.write(f"noisy_test_frr: {[round(num, 3) for num in noisy_test_frr]}\n\n")
        
        # 写入实际绘制的点坐标
        f.write("===== 实际绘制的ROC曲线点坐标 =====\n\n")
        
        f.write("clean dev点坐标 (FAPH, FRR):\n")
        for x, y in zip(clean_dev_points_x, clean_dev_points_y):
            f.write(f"({round(x, 4)}, {round(y, 4)})\n")
        f.write("\n")
        
        f.write("clean test点坐标 (FAPH, FRR):\n")
        for x, y in zip(clean_test_points_x, clean_test_points_y):
            f.write(f"({round(x, 4)}, {round(y, 4)})\n")
        f.write("\n")
        
        f.write("noisy dev点坐标 (FAPH, FRR):\n")
        for x, y in zip(noisy_dev_points_x, noisy_dev_points_y):
            f.write(f"({round(x, 4)}, {round(y, 4)})\n")
        f.write("\n")
        
        f.write("noisy test点坐标 (FAPH, FRR):\n")
        for x, y in zip(noisy_test_points_x, noisy_test_points_y):
            f.write(f"({round(x, 4)}, {round(y, 4)})\n")
            
    print(f"ROC data saved to {txt_save_path}")

if __name__ == "__main__":
    main()
